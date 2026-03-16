
import io
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber
import streamlit as st

st.set_page_config(page_title="PDF nóminas -> Holded", layout="wide")

TEMPLATE_HEADERS = [
    "Documento de identidad",
    "Nombre empleado",
    "Fecha dd/mm/yyyy",
    "Descripción",
    "Salario",
    "Salario Cuenta (640)",
    "Total S.S.",
    "Total S.S. Cuenta (476)",
    "Gasto S.S. Empresa",
    "Gasto S.S. Empresa Cuenta (642)",
    "IRPF",
    "IRPF Cuenta (4751)",
    "Tags separados por -",
    "Cantidad del pago",
    "Fecha de pago",
    "Cuenta de pago",
]

MONTHS_ABBR = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}


def extract_text(file_obj) -> str:
    with pdfplumber.open(file_obj) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def euro_to_float(value: str | None) -> float | None:
    if not value:
        return None
    value = value.replace(".", "").replace(",", ".")
    value = re.sub(r"[^0-9\.-]", "", value)
    return float(value) if value else None


def format_date(dt: datetime | None) -> str | None:
    return dt.strftime("%d/%m/%Y") if dt else None


def clean_name(name: str | None) -> str | None:
    if not name:
        return name
    name = re.sub(r"\s+(JEFE|COCIN.*|AUXILIAR.*|ADMINISTRATIVO.*)$", "", name).strip()
    name = re.sub(r"\s{2,}", " ", name)
    return name


def parse_model_sara(text: str, filename: str) -> dict:
    second_line = text.splitlines()[1].strip()
    employee_from_file = re.split(r"[-_]", Path(filename).stem)[0].strip().upper()

    employee = employee_from_file
    company = second_line[:-len(employee_from_file)].strip() if second_line.endswith(employee_from_file) else None

    nif = re.search(r"NIF:([A-Z0-9]+)", text)
    period = re.search(
        r"Periodo de Liquidación:del (\d{2})/(\d{2})/(\d{4}) al (\d{2})/(\d{2})/(\d{4})",
        text,
    )
    date = datetime(int(period.group(6)), int(period.group(5)), int(period.group(4))) if period else None

    salary = euro_to_float(re.search(r"A\. TOTAL DEVENGADO\s+([\d\.,]+)", text).group(1))
    total_ss = euro_to_float(re.search(r"1\. TOTAL APORTACIONES\s+([\d\.,]+)", text).group(1))
    irpf = euro_to_float(re.search(r"2\. I\.R\.P\.F\.\s+[\d,]+\s*%\s*([\d\.,]+)", text).group(1))
    net = euro_to_float(re.search(r"LIQUIDO TOTAL A PERCIBIR \(A - B\)\s+([\d\.,]+)", text).group(1))
    employer_ss = euro_to_float(re.search(r"Total aportación empresarial\s+([\d\.,]+)", text).group(1))

    return {
        "archivo": filename,
        "empresa": company,
        "empleado": clean_name(employee),
        "nif": nif.group(1) if nif else None,
        "fecha": date,
        "salario": salary,
        "total_ss": total_ss,
        "gasto_ss_empresa": employer_ss,
        "irpf": irpf,
        "neto": net,
        "modelo": "SARA/estándar",
    }


def parse_model_la_jaula(text: str, filename: str) -> dict:
    employee_line = None
    for line in text.splitlines():
        if "VALDAYO DOMINGUEZ" in line or "D.N.I." in line:
            employee_line = line

    worker_match = re.search(
        r"TRABAJADOR/A.*?\n([A-ZÁÉÍÓÚÑ ,.-]+?)\s+[A-Z]{3,}(?:\s+[A-Z]+)?\s+\d+\s+[A-Z]{3}\s+\d{2}\s+(\d{8}[A-Z])",
        text,
        re.S,
    )

    employee = worker_match.group(1).strip() if worker_match else None
    nif = worker_match.group(2) if worker_match else None

    if employee:
        employee = clean_name(employee.replace(" JEFE", "").strip())

    period = re.search(r"(\d{2}) ([A-Z]{3}) (\d{2}) a (\d{2}) ([A-Z]{3}) (\d{2})", text)
    date = datetime(2000 + int(period.group(6)), MONTHS_ABBR[period.group(5)], int(period.group(4))) if period else None

    salary_match = re.search(
        r"BASE I\.R\.P\.F\. T\. DEVENGADO T\. A DEDUCIR\s+[\d\.,]+\s+[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)\s+([\d\.,]+)",
        text,
    )
    salary = euro_to_float(salary_match.group(1)) if salary_match else None

    irpf = euro_to_float(re.search(r"TRIBUTACION I\.R\.P\.F\.\s*[\d,]+\s+([\d\.,]+)", text).group(1))
    net = euro_to_float(re.search(r"LIQUIDO A PERCIBIR\s+([\d\.,]+)", text).group(1))
    cost_company = euro_to_float(re.search(r"COSTE EMPRESA:\s*([\d\.,]+)", text).group(1))

    ss_parts = re.findall(
        r"COTIZACION (?:CONT\.COMU|MEI|FORMACION|DESEMPLEO)\s+[\d,]+\s+([\d\.,]+)",
        text,
    )
    total_ss = round(sum(euro_to_float(x) for x in ss_parts), 2) if ss_parts else None
    employer_ss = round(cost_company - salary, 2) if (cost_company is not None and salary is not None) else None

    return {
        "archivo": filename,
        "empresa": "LA JAULA DE PAPEL SL",
        "empleado": employee,
        "nif": nif,
        "fecha": date,
        "salario": salary,
        "total_ss": total_ss,
        "gasto_ss_empresa": employer_ss,
        "irpf": irpf,
        "neto": net,
        "modelo": "La Jaula",
    }


def parse_payroll(file_obj, filename: str) -> dict:
    text = extract_text(file_obj)
    if "Periodo de Liquidación:del" in text:
        return parse_model_sara(text, filename)
    return parse_model_la_jaula(text, filename)


def make_workbook(rows: list[dict], config: dict, template_bytes: bytes | None) -> bytes:
    if template_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb[wb.sheetnames[0]]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holded"
        ws.append(TEMPLATE_HEADERS)

    for row in rows:
        ws.append([
            row.get("nif"),
            row.get("empleado"),
            format_date(row.get("fecha")),
            f"Nómina {row.get('empleado') or ''} {row.get('fecha').strftime('%m/%Y') if row.get('fecha') else ''}".strip(),
            row.get("salario"),
            config["cuenta_640"],
            row.get("total_ss"),
            config["cuenta_476"],
            row.get("gasto_ss_empresa"),
            config["cuenta_642"],
            row.get("irpf"),
            config["cuenta_4751"],
            config["tags"],
            row.get("neto"),
            format_date(row.get("fecha")),
            config["cuenta_pago"],
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=14):
        for cell in row:
            cell.number_format = "0.00"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


st.title("Importar nóminas PDF a Holded")
st.caption("Carga uno o varios PDFs, detecta los dos modelos de nómina que me has pasado y genera el Excel listo para Holded.")

with st.sidebar:
    st.subheader("Configuración contable")
    cuenta_640 = st.text_input("Cuenta salario (640)", "64000000")
    cuenta_476 = st.text_input("Cuenta S.S. trabajador (476)", "47600000")
    cuenta_642 = st.text_input("Cuenta S.S. empresa (642)", "64200000")
    cuenta_4751 = st.text_input("Cuenta IRPF (4751)", "47510000")
    cuenta_pago = st.text_input("Cuenta de pago", "57200000")
    tags = st.text_input("Tags", "nominas")

template_file = st.file_uploader("Plantilla Excel de Holded", type=["xlsx"])
pdf_files = st.file_uploader("PDFs de nóminas", type=["pdf"], accept_multiple_files=True)

if st.button("Procesar", type="primary", disabled=not pdf_files):
    parsed_rows = []
    for pdf in pdf_files:
        pdf.seek(0)
        parsed_rows.append(parse_payroll(pdf, pdf.name))

    st.subheader("Datos extraídos")
    st.dataframe(parsed_rows, use_container_width=True)

    output = make_workbook(
        parsed_rows,
        {
            "cuenta_640": cuenta_640,
            "cuenta_476": cuenta_476,
            "cuenta_642": cuenta_642,
            "cuenta_4751": cuenta_4751,
            "cuenta_pago": cuenta_pago,
            "tags": tags,
        },
        template_file.getvalue() if template_file else None,
    )

    st.download_button(
        "Descargar Excel para Holded",
        data=output,
        file_name="nominas_holded.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.info("La columna 'Cantidad del pago' se rellena con el líquido a percibir de la nómina.")
